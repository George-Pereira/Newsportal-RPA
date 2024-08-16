from openpyxl import Workbook

class Excel:
    def __init__(self, filename:str) -> None:
        self.__workbook = Workbook()
        self.__filename = filename
        pass
    def getWorkbook(self):
        return self.__workbook
    def setWorkbook(self):
        self.__workbook = self
    def getFilename(self):
        return self.__filename
    def setFilename(self, filename:str):
        self.__filename = filename
